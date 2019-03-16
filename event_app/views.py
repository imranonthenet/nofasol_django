from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseRedirect, Http404
from .forms import EditForm, ImportDataForm
from .models import Event, EventField,EventBadgeCategory,BadgeCategory, ImportData, EventData
import os
from django.core.files.storage import FileSystemStorage
from django.urls import reverse
import openpyxl

# Create your views here.
def event_list(request):
    events = Event.objects.all()
    print('events=',events)
    return render(request, 'event_app/event_list.html', {'events':events})
    
def import_data(request, id):
    event = get_object_or_404(Event, id=id)

    if request.method == 'POST':
        form = ImportDataForm(request.POST, request.FILES)

        if form.is_valid():
            import_data = form.save()
            wb = openpyxl.load_workbook(import_data.import_file.path)
            sheets_list = wb.sheetnames

            sheet = wb[sheets_list[0]]
            eventData = EventData()
            eventData.event = event

            for row in range(2,sheet.max_row + 1):
                eventData.uniqueId = sheet.cell(row,1).value
                eventData.barcode = sheet.cell(row,2).value
                eventData.sno = sheet.cell(row,3).value
                eventData.title = sheet.cell(row,1).value
                eventData.firstName = sheet.cell(row,1).value
                eventData.middleName = sheet.cell(row,1).value
                eventData.lastName = sheet.cell(row,1).value
                eventData.fullName = sheet.cell(row,1).value
                eventData.jobTitle = sheet.cell(row,1).value
                eventData.department = sheet.cell(row,1).value
                eventData.companyName = sheet.cell(row,1).value
                eventData.mobile1 = sheet.cell(row,1).value
                eventData.mobile2 = sheet.cell(row,1).value
                eventData.tel1 = sheet.cell(row,1).value
                eventData.tel2 = sheet.cell(row,1).value
                eventData.fax = sheet.cell(row,1).value
                eventData.email = sheet.cell(row,1).value
                eventData.website = sheet.cell(row,1).value
                eventData.address1 = sheet.cell(row,1).value
                eventData.address2 = sheet.cell(row,1).value
                eventData.city = sheet.cell(row,1).value
                eventData.country = sheet.cell(row,1).value
                eventData.poBox = sheet.cell(row,1).value
                eventData.postalCode = sheet.cell(row,1).value
                eventData.badgeCategory = sheet.cell(row,1).value
                eventData.regType = 'Online' #sheet.cell(row,1).value
                eventData.regDate = sheet.cell(row,1).value
                eventData.badgePrintDate = sheet.cell(row,1).value
                eventData.modifiedDate = sheet.cell(row,1).value
                eventData.statusFlag = 'Did Not Attend' #sheet.cell(row,1).value
                eventData.backoffice = sheet.cell(row,1).value
                eventData.comment1 = sheet.cell(row,1).value
                eventData.comment2 = sheet.cell(row,1).value
                eventData.comment3 = sheet.cell(row,1).value
                eventData.comment4 = sheet.cell(row,1).value
                eventData.comment5 = sheet.cell(row,1).value
                eventData.comment6 = sheet.cell(row,1).value
                eventData.comment7 = sheet.cell(row,1).value
                eventData.comment8 = sheet.cell(row,1).value
                eventData.comment9 = sheet.cell(row,1).value
                eventData.comment10 = sheet.cell(row,1).value

                eventData.save()
            
            return redirect('event_app:event_list')

    else:
        form = ImportDataForm()

    return render(request, 'event_app/import_data.html', {'form':form, 'event_id':id})

def create_event(request):
    if request.method == 'POST':
        form = EditForm(request.POST, request.FILES)
        if form.is_valid():
            event = form.save()
            insert_event_fields_with_defaults(event)
            insert_badge_categories_with_defaults(event)

            return redirect('event_app:event_fields', id=event.id)

    else:
        form = EditForm()
    
    return render(request, 'event_app/create_event.html', {'form': form})

def edit(request, id):
    event = get_object_or_404(Event, id=id)

    if request.method == "POST":
        form = EditForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            event = form.save()
            return redirect('event_app:event_fields', id=event.id)

    else:
        form = EditForm(instance=event)

    return render(request, 'event_app/edit_event.html', {'form': form, 'event_id':id})


def create_event1(request):

    if request.method == 'POST':
        
        form = EditForm(request.POST, request.FILES)

        if form.is_valid():

            logo_file = request.FILES['event_logo']
            #fs = FileSystemStorage()
            #filename = fs.save(logo_file.name, logo_file)
            #uploaded_file_url = fs.url(filename)
            #request.session['event_logo'] = uploaded_file_url
            request.session['event_logo'] = '/static/' + logo_file.name
            handle_uploaded_file(logo_file)

            event = Event()
            event.event_name = form.cleaned_data['event_name']
            event.event_logo = logo_file.name
            event.event_from_date = form.cleaned_data['event_from_date']
            event.event_to_date = form.cleaned_data['event_to_date']
            event.setup_complete = False
            event.save()

            insert_event_fields_with_defaults(event)
 
           
            # return HttpResponseRedirect(reverse('event_app:event_fields', args=[event.id] ))
            return redirect('event_app:event_fields', id=event.id)
    else:
        form = EditForm()


    return render(request, 'event_app/create_event.html', {'form': form})


def edit1(request, id):
    event = get_object_or_404(Event, id=id)
    form = EditForm(
        {
            'event_name':event.event_name, 
            'event_logo':event.event_logo,
            'event_from_date':event.event_from_date,
            'event_to_date':event.event_to_date
        })

    if request.method == 'POST':
        
        form = EditForm(request.POST, request.FILES)

        if form.is_valid():
            print(form.cleaned_data['event_name'])
            print(form.cleaned_data['event_from_date'])
            print(form.cleaned_data['event_to_date'])

            logo_file = request.FILES['event_logo']
            #fs = FileSystemStorage()
            #filename = fs.save(logo_file.name, logo_file)
            #uploaded_file_url = fs.url(filename)
            #request.session['event_logo'] = uploaded_file_url
            request.session['event_logo'] = '/static/' + logo_file.name
            handle_uploaded_file(logo_file)

            event.event_name = form.cleaned_data['event_name']
            event.event_logo = logo_file.name
            event.event_from_date = form.cleaned_data['event_from_date']
            event.event_to_date = form.cleaned_data['event_to_date']
            event.setup_complete = False
            event.save()
 
           
            # return HttpResponseRedirect(reverse('event_app:event_fields', args=[event.id] ))
            return redirect('event_app:event_fields', id=id)



    return render(request, 'event_app/edit.html', {'form': form, 'event_id':id})


def event_fields(request, id):
    if request.method == 'POST':
        form_field_values = request.POST
        db_fields = EventField.objects.filter(event__id=id)

        print(form_field_values)

        for field in db_fields:
            field_label = form_field_values[f'{field.field_name}.field_label']


            is_mandatory = form_field_values.get(f'{field.field_name}.is_mandatory','') == 'on'
            show_in_search = form_field_values.get(f'{field.field_name}.show_in_search','') == 'on'
            include_in_search = form_field_values.get(f'{field.field_name}.include_in_search','') == 'on'
            show_in_register = form_field_values.get(f'{field.field_name}.show_in_register','') == 'on'
            show_in_print = form_field_values.get(f'{field.field_name}.show_in_print','') == 'on'
            column_in_excel = form_field_values[f'{field.field_name}.column_in_excel']

            event_field = EventField.objects.get(id=field.id)

            event_field.field_label = field_label
            event_field.is_mandatory = is_mandatory
            event_field.show_in_search = show_in_search
            event_field.include_in_search = include_in_search
            event_field.show_in_register = show_in_register
            event_field.show_in_print = show_in_print
            event_field.column_in_excel = column_in_excel

            event_field.save()

        
        return HttpResponseRedirect(reverse('event_app:badge_categories', args=[id]))

    else:

        db_fields = EventField.objects.filter(event__id=id)


    return render(request,'event_app/event_fields.html', {'fields':db_fields, 'event_id':id})

def badge_categories(request, id):
    if request.method == 'POST':
        form_field_values = request.POST
        db_fields = EventBadgeCategory.objects.filter(event__id=id)

        for field in db_fields:
            
            # this will set show_in_register to True, only if the value of 
            # `show_in_register` checkbox passed in the POST is on 

            show_in_register = request.POST.get(f'{field.badge_category_code}.show_in_register', '') == 'on'

            field.show_in_register = show_in_register
            field.save()


        return HttpResponseRedirect(reverse('event_app:badge_layout', kwargs={'id':id}) ) 

    else:
        badge_categories = EventBadgeCategory.objects.filter(event__id=id)
        return render(request,'event_app/badge_categories.html', {'fields':badge_categories, 'event_id':id  })

def badge_layout(request,id):
    show_barcode=False
    barcode_left = 0
    barcode_top = 0
    field_top_value = 0

    badge_fields = EventField.objects.filter(event__id=id,show_in_print=True)
    print(badge_fields)

    for index,field in enumerate(badge_fields):
        
        if field.top == 10:
            field.top = index * 20
        

        if field.field_name == 'barcode':
            show_barcode=True
            barcode_left = field.left
            barcode_top = field.top

    if request.method == 'POST':
        form_field_values = request.POST
        print(form_field_values)

        for field in badge_fields:
            print(field.field_name + form_field_values[f'{field.field_name}_top'])

            field.top = form_field_values[f'{field.field_name}_top']
            field.left = form_field_values[f'{field.field_name}_left']

            if field.field_name!='barcode':
                field.width = form_field_values[f'{field.field_name}_width']
                field.font_family = form_field_values[f'{field.field_name}_fontFamily']
                field.font_size = form_field_values[f'{field.field_name}_fontSize']
                field.font_weight = form_field_values[f'{field.field_name}_fontWeight']
                field.font_style = form_field_values[f'{field.field_name}_fontStyle']
                field.text_align = form_field_values[f'{field.field_name}_textAlign']
            
            field.save()
            event = Event.objects.get(id=id)
            event.setup_complete = True
            event.save()

        return redirect('event_app:event_list')
    
    return render(request,'event_app/badge_layout.html', {'fields':badge_fields, 'event_id':id, 'show_barcode':show_barcode,'barcode_left':barcode_left,'barcode_top':barcode_top} )

def handle_uploaded_file(f):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    upload_file = os.path.join(BASE_DIR,'static',f.name)

    with open(upload_file,'wb+') as file:
        for chunk in f.chunks():
            file.write(chunk)

def insert_badge_categories_with_defaults(event):


    EventBadgeCategory.objects.get_or_create(event=event, badge_category_code='VIP',badge_category_desc='VIP', show_in_register=True)
    EventBadgeCategory.objects.get_or_create(event=event, badge_category_code='Student',badge_category_desc='Student', show_in_register=False)


def insert_event_fields_with_defaults(event):

    EventField.objects.get_or_create(
        event = event,
        field_name = 'unique_id',
        field_type = 'text',
        field_label = 'Unique Id',
        is_mandatory = False,
        show_in_search = True,
        include_in_search = True,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'A',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'barcode',
        field_type = 'text',
        field_label = 'Barcode',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = True,
        column_in_excel = 'B',
        top = 10,
        left = 10
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'sno',
        field_type = 'text',
        field_label = 'SNo',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'C',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'title',
        field_type = 'text',
        field_label = 'Title',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'D',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'first_name',
        field_type = 'text',
        field_label = 'First Name',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'E',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'middle_name',
        field_type = 'text',
        field_label = 'Middle Name',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'F',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )


    EventField.objects.get_or_create(
        event = event,
        field_name = 'last_name',
        field_type = 'text',
        field_label = 'Last Name',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'G',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )


    EventField.objects.get_or_create(
        event = event,
        field_name = 'full_name',
        field_type = 'text',
        field_label = 'Full Name',
        is_mandatory = True,
        show_in_search = True,
        include_in_search = True,
        show_in_register = True,
        show_in_print = True,
        column_in_excel = 'H',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

  
    EventField.objects.get_or_create(
        event = event,
        field_name = 'job_title',
        field_type = 'text',
        field_label = 'Job Title',
        is_mandatory = True,
        show_in_search = True,
        include_in_search = True,
        show_in_register = True,
        show_in_print = False,
        column_in_excel = 'I',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

 
    EventField.objects.get_or_create(
        event = event,
        field_name = 'department',
        field_type = 'text',
        field_label = 'Department',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'J',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    
    EventField.objects.get_or_create(
        event = event,
        field_name = 'company_name',
        field_type = 'text',
        field_label = 'Last Name',
        is_mandatory = True,
        show_in_search = True,
        include_in_search = True,
        show_in_register = True,
        show_in_print = True,
        column_in_excel = 'K',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'mobile',
        field_type = 'text',
        field_label = 'Mobile',
        is_mandatory = True,
        show_in_search = True,
        include_in_search = True,
        show_in_register = True,
        show_in_print = False,
        column_in_excel = 'L',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'mobile2',
        field_type = 'text',
        field_label = 'Mobile 2',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'M',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'tel1',
        field_type = 'text',
        field_label = 'Tel 1',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'N',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'tel2',
        field_type = 'text',
        field_label = 'Tel 2',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'O',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'fax',
        field_type = 'text',
        field_label = 'Fax',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'P',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'email',
        field_type = 'text',
        field_label = 'Email',
        is_mandatory = True,
        show_in_search = True,
        include_in_search = True,
        show_in_register = True,
        show_in_print = False,
        column_in_excel = 'Q',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'website',
        field_type = 'text',
        field_label = 'Website',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'R',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'address1',
        field_type = 'text',
        field_label = 'Address 1',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'S',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )


    EventField.objects.get_or_create(
        event = event,
        field_name = 'address2',
        field_type = 'text',
        field_label = 'Address 2',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'T',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'city',
        field_type = 'text',
        field_label = 'City',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'U',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'country',
        field_type = 'text',
        field_label = 'Country',
        is_mandatory = True,
        show_in_search = True,
        include_in_search = True,
        show_in_register = True,
        show_in_print = False,
        column_in_excel = 'V',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'p_o_box',
        field_type = 'text',
        field_label = 'P.O.Box',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'W',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'postal_code',
        field_type = 'text',
        field_label = 'Postal Code',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'X',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'badge_category',
        field_type = 'text',
        field_label = 'Badge Category',
        is_mandatory = True,
        show_in_search = True,
        include_in_search = False,
        show_in_register = True,
        show_in_print = False,
        column_in_excel = 'Y',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'reg_type',
        field_type = 'text',
        field_label = 'Reg Type',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'Z',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'reg_date',
        field_type = 'text',
        field_label = 'Reg Date',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AA',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )


    EventField.objects.get_or_create(
        event = event,
        field_name = 'badge_print_date',
        field_type = 'text',
        field_label = 'Badge Print Date',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AB',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'modified_date',
        field_type = 'text',
        field_label = 'Modified Date',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AC',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'status_flag',
        field_type = 'text',
        field_label = 'Status Flag',
        is_mandatory = False,
        show_in_search = True,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AD',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'backoffice',
        field_type = 'text',
        field_label = 'Backoffice',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AE',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment1',
        field_type = 'text',
        field_label = 'Comment 1',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AF',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )


    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment2',
        field_type = 'text',
        field_label = 'Comment 2',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AG',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )

    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment3',
        field_type = 'text',
        field_label = 'Comment 3',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AH',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )    

    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment4',
        field_type = 'text',
        field_label = 'Comment 4',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AI',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )       

    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment5',
        field_type = 'text',
        field_label = 'Comment 5',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AJ',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )        

    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment6',
        field_type = 'text',
        field_label = 'Comment 6',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AK',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )        

    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment7',
        field_type = 'text',
        field_label = 'Comment 7',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AL',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )       

    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment8',
        field_type = 'text',
        field_label = 'Comment 8',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AM',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )       

    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment9',
        field_type = 'text',
        field_label = 'Comment 9',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AN',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )      

    EventField.objects.get_or_create(
        event = event,
        field_name = 'comment10',
        field_type = 'text',
        field_label = 'Comment 10',
        is_mandatory = False,
        show_in_search = False,
        include_in_search = False,
        show_in_register = False,
        show_in_print = False,
        column_in_excel = 'AO',
        top = 10,
        left = 10,
        width = 100,
        font_family = 'Calibri',
        font_size = 11,
        font_weight = 'normal',
        font_style = 'normal',
        text_align = 'left'
    )      